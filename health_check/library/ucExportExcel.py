#!/usr/bin/env  python3

import q
import os
from openpyxl import Workbook
from datetime import datetime
from ansible.module_utils.basic import *

class WriteExcel(object):

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "default"

    def set_sheetname(self, worksheet, sheename):
        worksheet.title = sheename

    def set_title(self, worksheet, title):
        worksheet.append(title)

    def append(self, worksheet, row):
        worksheet.append(row)

    def save(self, filepath):
        filedir = os.path.dirname(filepath)
        if not os.path.exists(filedir):
            os.mkdir(filedir)
        self.wb.save(filepath)

    def hosts(self, hostgroups, hostvars):
        hosts = []
        for hostgroup in hostgroups:
            for host in list(hostvars.values())[0]['groups'][hostgroup]:
                hosts.append(host)
        return hosts

    def result(self, host, hostvars, uclist):
        results = []
        if host in hostvars:
            hostvar = hostvars[host]
            keys = list(hostvar.keys())
            for k in keys:
                if k == 'res_user_info': 
                    user_info = hostvar[k]['user_info']
                    for k in list(user_info.keys()):
                        c_dict = user_info[k]
                        c_list = []
                        for k in uclist:
                            c_list.append(c_dict.get(k, None))
                        results.append(c_list)
        return results

    def walk_sheet(self, ws_name):
        ws = self.wb[ws_name]
        q(ws_name)
        result = []
        for r in range(1, ws.max_row+1):
            if r == 1:
                klist = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
                q(klist)
            else:
                vlist = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
                q(vlist)
                result.append(dict(zip(klist, vlist)))
        return result

    def set_ws_content(self, ws_name):
        self.ws_content = self.walk_sheet(ws_name)

    def item_result_sheet(self, item):
        ws = self.wb.create_sheet(item)
        title = ['host', 'name', item, 'datetime']
        self.append(ws, title)
        if self.ws_content:
            for c in self.ws_content:
                q(c)
                vlist = [c[k] for k in title]
                q(vlist)
                self.append(ws, vlist)

def write_sheet(hostgroups,hostvars,uclist,time,filepath):
    title = uclist.copy()
    title.insert(0, 'host')
    title.insert(len(title), 'datetime')
    q(title)
    excel = WriteExcel()
    excel.set_title(excel.ws, title)
    hosts = excel.hosts(hostgroups, hostvars)
    q(hosts)
    for host in hosts:
        q(host)
        results = excel.result(host, hostvars, uclist)
        for result in results:
            result.insert(0, host)
            q(title)
            result.insert(len(result), str(time))
            q(result)
            excel.append(excel.ws, result)
    excel.set_ws_content(excel.ws.title)

    for item in uclist:
        if item == 'name': continue
        excel.item_result_sheet(item)

    q(filepath)
    excel.save(filepath)


def main():
    module_args = dict(
            hostvars=dict(type='dict', required=True),
            hostgroups=dict(type='list', required=True),
            resultfile=dict(type='str', required=True),
            uclist=dict(type='list', required=True),
            )

    module = AnsibleModule(
            argument_spec=module_args,
            supports_check_mode=True,
            no_log=True,

            )

    result = dict(
            changed=False,
            message={}
            )

    hostvars = module.params['hostvars']
    hostgroups = module.params['hostgroups']
    resultfile = module.params['resultfile']
    uclist = module.params['uclist']

    q(uclist)
    time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    q(time)
    q(resultfile)
    q(hostgroups)
    write_sheet(hostgroups, hostvars, uclist, time, resultfile)

    result['changed'] = True
    module.exit_json(**result)

main()
