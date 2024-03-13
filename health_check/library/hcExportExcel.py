#!/usr/bin/env  python3

import q
import os
from openpyxl import Workbook,load_workbook
from datetime import datetime
from ansible.module_utils.basic import *

class WriteExcel(object):

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "default"
        self.ws_content = []
        self.items = []

    def set_sheetname(self, worksheet, sheename):
        worksheet.title = sheename

    def set_title(self, worksheet, title):
        worksheet.append(title)

    def append(self, worksheet, row):
        worksheet.append(row)

    def save(self, filepath):
        filedir = os.path.dirname(filepath)
        if not os.path.exists(filedir):
            os.mkdir(filedir)
        self.wb.save(filepath)

    def hosts(self, hostgroups, hostvars):
        hosts = []
        for hostgroup in hostgroups:
            for host in list(hostvars.values())[0]['groups'][hostgroup]:
                hosts.append(host)
        return hosts

    def set_items(self, item):
        self.items.append(item)

    def result(self, host, hostvars, hclist):
        results = []
        if host in hostvars:
            hostvar = hostvars[host]
            keys = list(hostvar.keys())
            for k in keys:
                k_result = []
                if k.startswith("res_"):
                    item = k.replace('res_', '')
                    q(item)
                    flag = item.split('_')[0]
                    q(flag)
                    if flag not in hclist:
                        continue
                    if item not in self.items:
                        self.set_items(item)
                    res = hostvar[k]
                    q(res)
                    if 'stdout' in res:
                        stdout = res['stdout'].strip()
                        q(stdout)
                    if 'rc' in res:
                        rc = res['rc']
                        if rc == 0:
                            rc = 'OK'
                        elif rc == 1:
                            rc = 'FAIL'
                        else:
                            rc = "ERROR"
                    k_result.append(item)
                    k_result.append(rc)
                    k_result.append(stdout)
                    results.append(k_result)
        return results

    def walk_sheet(self, ws_name):
        ws = self.wb[ws_name]
        q(ws_name)
        result = []
        for r in range(1, ws.max_row+1):
            if r == 1:
                klist = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
                q(klist)
            else:
                vlist = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
                q(vlist)
                result.append(dict(zip(klist, vlist)))
        return result

    def set_ws_content(self, ws_name):
        self.ws_content = self.walk_sheet(ws_name)

    def row_filter(self, item):
        item_content = []
        if self.ws_content:
            for line in self.ws_content:
                if line['item'] == item:
                    item_content.append(line)
        q(item_content)
        return item_content

    def item_result_sheet(self, item, title):
        ws= self.wb.create_sheet(item)
        self.append(ws, title)
        item_content = self.row_filter(item)
        for ic in item_content:
            vlist = [ic[k] for k in title]
            q(vlist)
            self.append(ws, vlist)

def write_sheet(hostgroups,hostvars,hclist,time,filepath):
    title = ['host', 'item', 'rc', 'stdout', 'datetime']
    excel = WriteExcel()
    excel.set_title(excel.ws, title)
    hosts = excel.hosts(hostgroups, hostvars)
    q(hosts)
    for host in hosts:
        q(host)
        results = excel.result(host, hostvars, hclist)
        for result in results:
            result.insert(0, host)
            result.insert(len(result), str(time))
            q(result)
            excel.append(excel.ws, result)
    excel.set_ws_content(excel.ws.title)
    q(excel.ws_content)
    q(excel.items)
    for item in excel.items:
        q(item)
        excel.item_result_sheet(item, title)

    q(filepath)
    excel.save(filepath)


def main():
    module_args = dict(
            hostvars=dict(type='dict', required=True),
            hostgroups=dict(type='list', required=True),
            resultfile=dict(type='str', required=True),
            hclist=dict(type='list', required=True),
            )

    module = AnsibleModule(
            argument_spec=module_args,
            supports_check_mode=True,
            no_log=True,

            )

    result = dict(
            changed=False,
            message={}
            )

    hostvars = module.params['hostvars']
    hostgroups = module.params['hostgroups']
    resultfile = module.params['resultfile']
    hclist = module.params['hclist']

    q(hclist)
    time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    q(time)
    write_sheet(hostgroups, hostvars, hclist, time, resultfile)
    #q(hostvars)

    result['changed'] = True
    module.exit_json(**result)

main()
