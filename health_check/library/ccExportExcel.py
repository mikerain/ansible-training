#!/usr/bin/env  python

import q
import os
from openpyxl import Workbook,load_workbook
from datetime import datetime
from ansible.module_utils.basic import *

class WriteExcel(object):

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "default"
        self.ws_content = []
        self.items = []

    def set_sheetname(self, worksheet, sheename):
        worksheet.title = sheename

    def set_title(self, worksheet, title):
        worksheet.append(title)

    def append(self, worksheet, row):
        worksheet.append(row)

    def save(self, filepath):
        filedir = os.path.dirname(filepath)
        if not os.path.exists(filedir):
            os.mkdir(filedir)
        self.wb.save(filepath)

    def hosts(self, hostgroups, hostvars):
        hosts = []
        for hostgroup in hostgroups:
            for host in list(hostvars.values())[0]['groups'][hostgroup]:
                hosts.append(host)
        return hosts

    def set_items(self, item):
        self.items.append(item)

    def result(self, host, hostvars, hclist):
        results = []
        if host in hostvars:
            hostvar = hostvars[host]
            keys = list(hostvar.keys())
            for k in keys:
                k_result = []
                if k.startswith("res_"):
                    item = k.replace('res_', '')
                    q(item)
                    flag = item.split('_')[0]
                    q(flag)
                    if flag not in hclist:
                        continue
                    if item not in self.items:
                        self.set_items(item)
                    res = hostvar[k]
                    q(res)
                    if 'stdout' in res:
                        stdout = res['stdout'].strip()
                        q(stdout)
                    if 'rc' in res:
                        rc = res['rc']
                        if rc == 0:
                            rc = 'OK'
                        elif rc == 1:
                            rc = 'FAIL'
                        else:
                            rc = "ERROR"
                    k_result.append(item)
                    k_result.append(rc)
                    k_result.append(stdout)
                    results.append(k_result)
        return results

    def walk_sheet(self, ws_name):
        ws = self.wb[ws_name]
        q(ws_name)
        result = []
        for r in range(1, ws.max_row+1):
            if r == 1:
                klist = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
                q(klist)
            else:
                vlist = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
                q(vlist)
                result.append(dict(zip(klist, vlist)))
        return result

    def set_ws_content(self, ws_name):
        self.ws_content = self.walk_sheet(ws_name)

    def row_filter(self, item):
        item_content = []
        if self.ws_content:
            for line in self.ws_content:
                if line['item'] == item:
                    item_content.append(line)
        q(item_content)
        return item_content

    def item_result_sheet(self, item, title):
        ws= self.wb.create_sheet(item)
        self.append(ws, title)
        item_content = self.row_filter(item)
        for ic in item_content:
            vlist = [ic[k] for k in title]
            q(vlist)
            self.append(ws, vlist)

    def compare_type(self, cmp_dict):
        q(cmp_dict)

        for item,target in list(cmp_dict.items()):
            res = self.row_filter(item)
            gt = []
            lt = []
            for line in res:
                real = line.get('stdout')
                host = line.get('host')
                q(host)
                q(item)
                q(target)
                q(real)
                content ="%s[%s];"%(host,real)
                if float(real) > float(target):
                    gt.append(content)
                else:
                    lt.append(content)
            q(gt)
            q(lt)
            row = []
            row.insert(0, item)
            row.insert(1, target)
            row.insert(2, ('\n').join(gt))
            row.insert(3, ('\n').join(lt))
            q(row)
            ws = self.wb.create_sheet(item)
            title = ['item', 'target', 'great_than target', 'less_then target']
            self.append(ws, title)
            self.append(ws, row)

    def equal_type(self, eq_dict):
        q(eq_dict)
        for item,target in list(eq_dict.items()):
            res = self.row_filter(item)
            eq = []
            not_eq = []
            for line in res:
                real = line.get('stdout')
                host = line.get('host')
                q(host)
                q(item)
                q(target)
                q(real)
                content ="%s[%s];"%(host,real)
                q(content)
                if real == target:
                    eq.append(content)
                else:
                    not_eq.append(content)
            q(eq)            
            q(not_eq)
            row = []
            row.insert(0, item)
            row.insert(1, target)
            row.insert(2, ('\n').join(eq))
            row.insert(3, ('\n').join(not_eq))
            q(row)
            ws = self.wb.create_sheet(item)
            title = ['item', 'target', 'equal_to target', 'not_equal_to target']
            self.append(ws, title)
            self.append(ws, row)

    def exists_type(self, ex_dict):
        q(ex_dict)
        for item,target in list(ex_dict.items()):
            res = self.row_filter(item)
            exists = []
            not_exists = []
            for line in res:
                real = line.get('stdout')
                host = line.get('host')
                q(host)
                q(item)
                q(target)
                q(real)
                content ="%s[%s];"%(host,real)
                if target in real:
                    exists.append(content)
                else:
                    not_exists.append(content)
            q(exists)
            q(not_exists)
            row = []
            row.insert(0, item)
            row.insert(1, target)
            row.insert(2, ('\n').join(exists))
            row.insert(3, ('\n').join(not_exists))
            q(row)
            ws = self.wb.create_sheet(item)
            title = ['item', 'target', 'target exists', 'target not exists']
            self.append(ws, title)
            self.append(ws, row)

    def disk_type(self, disk_dict):
        q(disk_dict)
        for item,target in list(disk_dict.items()):
            res = self.row_filter(item)
            gt = []
            lt = []
            for line in res:
                real = line.get('stdout')
                host = line.get('host')
                q(host)
                q(item)
                q(target)
                q(real)
                for r in real.split('\n'):
                    disk, usage = r.strip().split(':')
                    q(disk)
                    q(usage)
                    content = "%s.%s[%s]"%(host,disk,usage.strip())
                    q(content)
                    if float(usage.replace('%','')) > float(target):
                        gt.append(content)
                    else:
                        lt.append(content)
            q(gt)
            q(lt)
            row = []
            row.insert(0, item)
            row.insert(1, target)
            row.insert(2, ('\n').join(gt))
            row.insert(3, ('\n').join(lt))
            q(row)
            ws = self.wb.create_sheet(item)
            title = ['item', 'target', 'great_than target', 'less_then target']
            self.append(ws, title)
            self.append(ws, row)

def write_sheet(hostgroups,hostvars,hclist,ccdict,time,filepath):
    title = ['host', 'item', 'rc', 'stdout', 'datetime']
    excel = WriteExcel()
    excel.set_title(excel.ws, title)
    hosts = excel.hosts(hostgroups, hostvars)
    q(hosts)
    for host in hosts:
        q(host)
        results = excel.result(host, hostvars, hclist)
        for result in results:
            result.insert(0, host)
            result.insert(len(result), str(time))
            q(result)
            excel.append(excel.ws, result)
    excel.set_ws_content(excel.ws.title)
    q(excel.ws_content)
    q(excel.items)
    #q(ccdict)
    for k,v in list(ccdict.items()):
        if k == 'compare_type': excel.compare_type(v)
        elif k == 'equal_type': excel.equal_type(v)
        elif k == 'exists_type': excel.exists_type(v)
        elif k == 'disk_type': excel.disk_type(v)
            
    q(filepath)
    excel.save(filepath)


def main():
    module_args = dict(
            hostvars=dict(type='dict', required=True),
            hostgroups=dict(type='list', required=True),
            resultfile=dict(type='str', required=True),
            hclist=dict(type='list', required=True),
            ccdict=dict(type='dict', required=True),
            )

    module = AnsibleModule(
            argument_spec=module_args,
            supports_check_mode=True,
            no_log=True,

            )

    result = dict(
            changed=False,
            message={}
            )

    hostvars = module.params['hostvars']
    hostgroups = module.params['hostgroups']
    resultfile = module.params['resultfile']
    hclist = module.params['hclist']
    ccdict = module.params['ccdict']

    q(hclist)
    time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    q(time)
    write_sheet(hostgroups, hostvars, hclist, ccdict, time, resultfile)

    result['changed'] = True
    module.exit_json(**result)

main()
